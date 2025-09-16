import re
import json
from openpyxl import load_workbook

class Client:
    def __init__(self, surname: str, name: str, phone: str,
                   patronymic: str = "", passport: str = None, 
                   email: str = None, comment: str = ""):
          self._surname = self.validate_fio(surname, "Фамилия")
          self._name = self.validate_fio(name, "Имя")
          self._patronymic = self.validate_fio(patronymic, "Отчество") if patronymic else ""
          self._passport = self.validate_passport(passport)
          self._phone = self.validate_phone(phone)
          self._email = self.validate_email(email) if email else None
          self._comment = comment or ""

    @staticmethod
    def validate_required(value, field_name: str):
        if not value or not str(value).strip():
            raise ValueError(f"Поле '{field_name}' не может быть пустым!")
        return value

    @staticmethod
    def validate_fio(value: str, field_name: str):
        value = Client.validate_required(value, field_name)
        pattern = r"^[А-Яа-яЁёA-Za-z]+$"
        if not re.match(pattern, value):
            raise ValueError(f"{field_name} должно содержать только буквы!")
        return value

    @staticmethod
    def validate_passport(passport: str):
        if not passport:
            return None
        passport_clean = str(passport).replace(" ", "")
        pattern = r"^\d{10}$"
        if not re.match(pattern, passport_clean):
            raise ValueError("Паспорт должен состоять ровно из 10 цифр!")
        return passport_clean

    @staticmethod
    def validate_phone(phone: str):
        phone = Client.validate_required(phone, "Телефон")
        phone = str(phone)
        pattern = r"^\+?\d{7,11}$"
        if not re.match(pattern, phone):
            raise ValueError("Телефон должен содержать только цифры и '+' (от 7 до 11 символов)!")
        return phone

    @staticmethod
    def validate_email(email: str):
        if not email:
            return None
        pattern = r"^(?!\.)[A-Za-z0-9](?:[A-Za-z0-9._%+-]*)@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$"
        if ".." in email:
            raise ValueError("Email не может содержать подряд идущие точки!")
        if not re.match(pattern, email):
            raise ValueError("Некорректный формат email!")
        return email
      
    @property
    def surname(self): 
        return self._surname

    @surname.setter
    def surname(self, value: str):
        self._surname = self.validate_fio(value, "Фамилия")

    @property
    def name(self): 
        return self._name

    @name.setter
    def name(self, value: str):
        self._name = self.validate_fio(value, "Имя")

    @property
    def patronymic(self): 
        return self._patronymic

    @patronymic.setter
    def patronymic(self, value: str):
        self._patronymic = self.validate_fio(value, "Отчество") if value else ""

    @property
    def passport(self): 
        return self._passport

    @passport.setter
    def passport(self, value: str):
        self._passport = self.validate_passport(value)

    @property
    def phone(self): 
        return self._phone

    @phone.setter
    def phone(self, value: str):
        self._phone = self.validate_phone(value)

    @property
    def email(self): 
        return self._email

    @email.setter
    def email(self, value: str):
        self._email = self.validate_email(value) if value else None

    @property
    def comment(self): 
        return self._comment

    @comment.setter
    def comment(self, value: str):
        self._comment = value or ""

    @classmethod
    def from_string(cls, line: str, delimiter=";"):
        parts = line.strip().split(delimiter)
        if len(parts) < 3:
            raise ValueError("Строка должна содержать минимум фамилию, имя и телефон!")
        return cls(
            surname=parts[0],
            name=parts[1],
            patronymic=parts[2],
            passport=parts[3],
            phone=parts[4],
            email=parts[5] if len(parts) > 5 else None,
            comment=parts[6] if len(parts) > 6 else ""
        )

    @classmethod
    def from_txt(cls, filepath: str, delimiter=";"):
        with open(filepath, encoding="utf-8") as f:
            line = f.readline()
        return cls.from_string(line, delimiter)

    @classmethod
    def from_dict(cls, data: dict):
        required_fields = ["surname", "name", "phone"]
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                raise ValueError(f"Обязательное поле '{field}' отсутствует или пустое!")

        return cls(
            surname=data["surname"],
            name=data["name"],
            patronymic=data.get("patronymic", ""),
            passport=data.get("passport"),
            phone=data["phone"],
            email=data.get("email"),
            comment=data.get("comment", "")
        )

    @classmethod
    def from_excel(cls, filepath: str, sheet=0, row=2):
        wb = load_workbook(filepath)
        ws = wb[wb.sheetnames[sheet]]
        data = {
            "surname": ws.cell(row=row, column=1).value,
            "name": ws.cell(row=row, column=2).value,
            "patronymic": ws.cell(row=row, column=3).value,
            "passport": ws.cell(row=row, column=4).value,
            "phone": ws.cell(row=row, column=5).value,
            "email": ws.cell(row=row, column=6).value,
            "comment": ws.cell(row=row, column=7).value,
        }
        return cls.from_dict(data)

    @classmethod
    def from_json(cls, json_str: str):
        return cls.from_dict(json.loads(json_str))
